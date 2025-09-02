from flask import Flask, render_template, request, redirect, url_for, session, flash, Response, abort 
from flask import Flask, jsonify  # Add jsonify and session if missing
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.ext.hybrid import hybrid_property
from flask import make_response
from flask import send_file  # Add this with your other Flask imports
from flask import g
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from io import BytesIO
from urllib.parse import unquote
import sys  # Add this with other imports
from datetime import datetime, timezone, timedelta  # Add timezone here
from sqlalchemy import text
from sqlalchemy import String
from werkzeug.utils import secure_filename
from fpdf import FPDF
from collections import defaultdict
import csv
import sqlite3
import logging
from dotenv import load_dotenv
from functools import wraps
import os
import io
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from flask_migrate import Migrate
from reportlab.pdfgen import canvas
from reportlab.lib import utils
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
import json
from flask import jsonify
import re
import pandas as pd
import numpy as np
from sqlalchemy import func, and_, or_
import pytz
import requests
import ssl
import certifi
from urllib3.util.ssl_ import create_urllib3_context
from decimal import Decimal
import calendar  



# Set up logging
logger = logging.getLogger(__name__)

# Load the environment variables from the .env file
load_dotenv(dotenv_path="env.env")



# Load the environment variables from the .env file
# Get Google Maps API key
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")


# Create Flask app first (needed for app.logger)
app = Flask(__name__)

# Folder to store uploaded images (ensure this folder exists)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

# Make sure the 'uploads' folder exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Check if file is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

# Set the database URI to point to the instance folder
# Set the database URI correctly
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'instance', 'rrtables.db')
print("Database Path:", db_path)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv("FLASK_SECRET_KEY", "your_secret_key")
app.secret_key = 'your-secret-key-here'

print("Flask Secret Key:", app.config['SECRET_KEY'])

# Initialize the database
# Set up logging
logging.basicConfig(level=logging.INFO)

# Initialize extensions
db = SQLAlchemy(app)
migrate = Migrate(app, db)  # Add this line

login_manager = LoginManager()
login_manager.init_app(app)


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            flash('Please log in to access this page.', 'danger')
            return redirect(url_for('login'))
        if current_user.role != 'admin':
            flash('You do not have permission to access this page.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function
# Add these new before_request handlers here:
@app.before_request
def before_request():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(minutes=30)

# Configure upload folder
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)



class ProductionSchedule(db.Model):
    __tablename__ = 'production_schedule'
    
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    shift = db.Column(db.String(20), nullable=False)  # morning/afternoon/evening
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship to Product
    product = db.relationship('Product', backref='scheduled_productions')

class Distributor(db.Model):
    __tablename__ = 'distributors'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    contact_name = db.Column(db.String(100))
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    address = db.Column(db.Text)
    payment_terms = db.Column(db.Text)
    tax_id = db.Column(db.String(50))
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    purchase_orders = db.relationship('PurchaseOrder', backref='distributor', lazy=True)

class PurchaseOrder(db.Model):
    __tablename__ = 'purchase_orders'
    id = db.Column(db.Integer, primary_key=True)
    po_number = db.Column(db.String(20), unique=True, nullable=False)
    distributor_id = db.Column(db.Integer, db.ForeignKey('distributors.id'), nullable=False)
    order_date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    expected_delivery_date = db.Column(db.Date)
    status = db.Column(db.String(20), nullable=False, default='draft')
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    items = db.relationship('POItem', backref='purchase_order', lazy=True, cascade="all, delete-orphan")
    
    @property
    def total_amount(self):
        return sum(item.line_total for item in self.items)
    
    @property
    def total_amount(self):
        return sum(item.line_total for item in self.items)

class POItem(db.Model):
    __tablename__ = 'po_items'
    id = db.Column(db.Integer, primary_key=True)
    po_id = db.Column(db.Integer, db.ForeignKey('purchase_orders.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=True)
    ingredient_id = db.Column(db.Integer, db.ForeignKey('ingredients.id'), nullable=True)
    description = db.Column(db.String(255), nullable=True)  # For manual entries
    quantity = db.Column(db.Float, nullable=False)
    unit_price = db.Column(db.Float, nullable=False)
    unit_of_measure = db.Column(db.String(20), nullable=False, default='unit')
    received_quantity = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default='pending')
    
    # Relationships
    product = db.relationship('Product')
    ingredient = db.relationship('Ingredient')
    
    @property
    def line_total(self):
        return self.quantity * self.unit_price

# User Model
class User(db.Model, UserMixin):
    __tablename__ = 'user'
    user_id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False, unique=True)
    email = db.Column(db.String(100), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), nullable=False, default='employee')
    
    # NO relationship definition here - using backref from PunchRecord

    # Required methods for password handling
    def set_password(self, password):
        self.password = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password, password)

    def get_id(self):
        return str(self.user_id)

    def __repr__(self):
        return f'<User {self.username}>'


class ProductionRecord(db.Model):
    __tablename__ = 'production_records'
    id = db.Column(db.Integer, primary_key=True)
    punch_id = db.Column(db.Integer, db.ForeignKey('punch_record.punch_id'))
    user_id = db.Column(db.Integer, db.ForeignKey('user.user_id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'), nullable=True)
    product_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Float, nullable=True)
    lot_number = db.Column(db.String(50), nullable=True)
    production_date = db.Column(db.Date, nullable=False, default=datetime.utcnow().date)
    recorded_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    
    # Add relationship to User
    user = db.relationship('User', backref='production_records')
    
    # Add relationship to Product if needed
    product = db.relationship('Product', backref='production_records')

class Inventory(db.Model):
    __tablename__ = 'inventory'
    id = db.Column(db.Integer, primary_key=True)
    product_name = db.Column(db.String(100), nullable=False, unique=True)
    quantity = db.Column(db.Float, nullable=False, default=0.0)
    price = db.Column(db.Float, nullable=False, default=0.0)
    last_updated = db.Column(db.DateTime, default=datetime.utcnow)
class Ingredient(db.Model):
    __tablename__ = 'ingredients'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    unit_weight = db.Column(db.Float)
    unit_cost = db.Column(db.Float)
    # Add any other fields you need


class Product(db.Model):
    __tablename__ = 'products'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    selling_price = db.Column(db.Float, default=0.0)
    direct_costs_total = db.Column(db.Float, default=0.0)  # Remove the 'direct_costs' mapping
    indirect_costs = db.Column(db.Float, default=0.0)
    profit = db.Column(db.Float, default=0.0)
    
    @hybrid_property
    def direct_costs(self):
        return self.direct_costs_total
    
    @direct_costs.setter
    def direct_costs(self, value):
        self.direct_costs_total = value
    
    @direct_costs.expression
    def direct_costs(cls):
        return cls.direct_costs_total
    

class DirectCost(db.Model):
    __tablename__ = 'direct_costs'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    cost_type = db.Column(db.String(50))
    amount = db.Column(db.Float)
    date_recorded = db.Column(db.Date)
    product = db.relationship('Product', backref='cost_records')  # Unique backref name

class CostAllocationMethod(db.Model):
    __tablename__ = 'cost_allocation_methods'
    id = db.Column(db.Integer, primary_key=True)
    cost_type = db.Column(db.String(50))
    allocation_method = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class CostDriver(db.Model):
    __tablename__ = 'cost_drivers'
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('products.id'))
    date_recorded = db.Column(db.Date)
    labor_hours = db.Column(db.Float)
    machine_hours = db.Column(db.Float)
    production_volume = db.Column(db.Integer)
    revenue_share = db.Column(db.Float)

class Asset(db.Model):
    __tablename__ = 'assets'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    value = db.Column(db.Float)
    purchase_date = db.Column(db.Date)
    depreciation_rate = db.Column(db.Float)

class Recipe(db.Model):
    __tablename__ = 'recipes'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    labour_cost = db.Column(db.Float, nullable=False)
    ingredients = db.relationship('RecipeIngredient', backref='recipe', lazy=True)

class RecipeIngredient(db.Model):
    __tablename__ = 'recipe_ingredients'
    id = db.Column(db.Integer, primary_key=True)
    recipe_id = db.Column(db.Integer, db.ForeignKey('recipes.id'))
    ingredient_name = db.Column(db.String(100), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(20), nullable=False)






def initialize_database():
    """Safe database initialization that preserves admin users"""
    with app.app_context():
        # Get existing admin users before dropping tables
        admins = []
        try:
            admins = User.query.filter_by(role='admin').all()
            admins = [{'username': a.username, 'email': a.email, 'password': a.password} 
                     for a in admins]
        except:
            pass
        
        # Create fresh tables
        db.drop_all()
        db.create_all()
        
        # Restore admin users
        for admin in admins:
            if not User.query.filter_by(username=admin['username']).first():
                user = User(
                    username=admin['username'],
                    email=admin['email'],
                    password=admin['password'],
                    role='admin'
                )
                db.session.add(user)
        db.session.commit()
class ProjectCalendar(db.Model):
    __tablename__ = 'project_calendar'
    id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(100), nullable=False)
    date = db.Column(db.Date, nullable=False)
    start_time = db.Column(db.Time)
    end_time = db.Column(db.Time)
    team_members = db.Column(db.Text)  # Comma-separated list of usernames
    description = db.Column(db.Text)
    materials_needed = db.Column(db.Text)  # JSON string of materials requested
    status = db.Column(db.String(20), default='scheduled')  # scheduled, in-progress, completed
    created_by = db.Column(db.Integer, db.ForeignKey('user.user_id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationship
    creator = db.relationship('User', backref='created_events')

class Project(db.Model):
    __tablename__ = 'projects'
    project_id = db.Column(db.Integer, primary_key=True)
    project_name = db.Column(db.String(100), nullable=False, unique=True)
    client_name = db.Column(db.String(100))
    address = db.Column(db.Text)
    start_date = db.Column(db.Date)
    estimated_end_date = db.Column(db.Date)
    status = db.Column(db.String(20), default='active')
    budget = db.Column(db.Float)
    notes = db.Column(db.Text)
    
    # REMOVE the relationship to avoid circular reference
    # punch_records = db.relationship('PunchRecord', lazy=True)


class PunchRecord(db.Model):
    __tablename__ = 'punch_record'
    punch_id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.user_id'), nullable=False)
    punch_in_time = db.Column(db.DateTime, nullable=False)
    punch_out_time = db.Column(db.DateTime)
    total_hours_worked = db.Column(db.Float)
    date = db.Column(db.Date, nullable=False, default=datetime.utcnow)
    is_manual = db.Column(db.Boolean, default=False, nullable=False)
    project_name = db.Column(db.String(100))
    project_id = db.Column(db.Integer, db.ForeignKey('projects.project_id'))  # Foreign key
    task_description = db.Column(db.Text)
    progress_photo = db.Column(db.String(255))
    progress_photo_out = db.Column(db.String(255))
    progress_notes = db.Column(db.Text)
    latitude = db.Column(db.Float)
    longitude = db.Column(db.Float)
    location_address = db.Column(db.Text)
    latitude_out = db.Column(db.Float)
    longitude_out = db.Column(db.Float)
    location_address_out = db.Column(db.Text)
    
    # Relationship to user
    user = db.relationship('User', backref=db.backref('punch_records', lazy=True))
    # REMOVE the project relationship to avoid circular reference
    # project = db.relationship('Project', backref=db.backref('punch_records', lazy=True))

@app.cli.command('reset-db')
def reset_db_command():
    """Reset database while preserving structure and admin users"""
    with app.app_context():
        # Backup admin users
        admins = User.query.filter_by(role='admin').all()
        admin_data = [{
            'username': a.username,
            'email': a.email,
            'password': a.password,
            'role': 'admin'
        } for a in admins]
        
        # Get current schema version
        from flask_migrate import current
        try:
            current_rev = current()
        except:
            current_rev = None
        
        # Reset database
        db.drop_all()
        db.create_all()
        
        # Restore admins
        for admin in admin_data:
            if not User.query.filter_by(username=admin['username']).first():
                user = User(**admin)
                db.session.add(user)
        
        # Restore schema version if exists
        if current_rev:
            from flask_migrate import stamp
            stamp(revision=current_rev)
        
        db.session.commit()
    print("Database reset complete - admins and schema version preserved")

migrate = Migrate(app, db)

    
#  Role-Based Access Control: Admin and Employee
@app.route('/update_role/<int:user_id>', methods=['POST'])
@login_required
def update_role(user_id):
    if current_user.role != 'admin':
        flash('Access denied: Admins only', 'danger')
        return redirect(url_for('dashboard'))
    
    user = User.query.get(user_id)
    if not user:
        flash('User not found', 'danger')
        return redirect(url_for('dashboard'))
    
    new_role = request.form.get('role')
    if new_role not in ['admin', 'employee']:
        flash('Invalid role', 'danger')
        return redirect(url_for('dashboard'))
    
    user.role = new_role
    db.session.commit()
    flash(f'User {user.name} role updated to {new_role}', 'success')
    return redirect(url_for('dashboard'))

@app.route('/test_db')
def test_db():
    try:
        conn = sqlite3.connect(db_path)  # Use db_path here
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        conn.close()
        return f"Tables in database: {tables}"
    except Exception as e:
        return f"Error: {e}"

@app.route('/report')
@login_required
@admin_required
def report():
    users = User.query.all()
    user_records = {}
    
    local_tz = pytz.timezone('America/Toronto')
    
    # Get all unique projects for filter
    all_projects = set()
    total_hours_all = 0
    total_records = 0
    active_punches = 0
    project_hours = defaultdict(float)
    
    for user in users:
        records = PunchRecord.query.filter_by(user_id=user.user_id).order_by(PunchRecord.date.desc()).all()
        total_hours = sum(record.total_hours_worked or 0 for record in records)
        total_hours_all += total_hours
        total_records += len(records)
        
        formatted_records = []
        for record in records:
            # Count active punches
            if record.punch_out_time is None:
                active_punches += 1
            
            # Track project hours
            project_name = record.project_name or 'No Project'
            project_hours[project_name] += record.total_hours_worked or 0
            all_projects.add(project_name)
            
            local_punch_in = record.punch_in_time.replace(tzinfo=timezone.utc).astimezone(local_tz) if record.punch_in_time else None
            local_punch_out = record.punch_out_time.replace(tzinfo=timezone.utc).astimezone(local_tz) if record.punch_out_time else None
            
            # Get project name from Project table if available
            if record.project_id:
                project = Project.query.get(record.project_id)
                if project:
                    project_name = project.project_name
            
            formatted_records.append({
                'punch_id': record.punch_id,
                'date': record.date,
                'project_name': project_name,
                'task_description': record.task_description,
                'punch_in_time': local_punch_in,
                'punch_out_time': local_punch_out,
                'total_hours_worked': record.total_hours_worked,
                'is_manual': record.is_manual,
                'latitude': record.latitude,
                'longitude': record.longitude,
                'location_address': record.location_address,
                'progress_photo': record.progress_photo,
                'progress_photo_out': record.progress_photo_out,
                'latitude_out': record.latitude_out,
                'longitude_out': record.longitude_out,
                'location_address_out': record.location_address_out
            })
        
        user_records[user.user_id] = {
            'info': {'username': user.username, 'role': user.role},
            'records': formatted_records,
            'total_hours': total_hours
        }
    
    return render_template('report.html', 
                         user_records=user_records,
                         all_projects=sorted(all_projects),
                         total_employees=len(users),
                         total_hours_all=total_hours_all,
                         total_records=total_records,
                         active_punches=active_punches,
                         project_hours=dict(project_hours))

@app.route('/edit_record/<int:record_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_record(record_id):
    if current_user.role != 'admin':
        flash('Access denied: Admins only', 'danger')
        return redirect(url_for('dashboard'))

    record = PunchRecord.query.get_or_404(record_id)

    if request.method == 'POST':
        punch_in_time = request.form.get('punch_in_time')
        punch_out_time = request.form.get('punch_out_time')
        project_name = request.form.get('project_name')
        task_description = request.form.get('task_description')

        if punch_in_time:
            record.punch_in_time = datetime.strptime(punch_in_time, '%Y-%m-%dT%H:%M')

        if punch_out_time and punch_out_time.strip():
            record.punch_out_time = datetime.strptime(punch_out_time, '%Y-%m-%dT%H:%M')
        else:
            record.punch_out_time = None

        if project_name:
            record.project_name = project_name
            
        if task_description:
            record.task_description = task_description

        # Recalculate hours if both times are present
        if record.punch_in_time and record.punch_out_time:
            time_worked = record.punch_out_time - record.punch_in_time
            record.total_hours_worked = round(time_worked.total_seconds() / 3600, 2)

        db.session.commit()
        flash('Record updated successfully!', 'success')
        return redirect(url_for('report'))  

    return render_template('edit_record.html', record=record)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.route('/delete_record/<int:record_id>', methods=['POST'])
@login_required
@admin_required
def delete_record(record_id):
    record = PunchRecord.query.get_or_404(record_id)
    db.session.delete(record)
    db.session.commit()
    flash('Record deleted successfully', 'success')
    return redirect(url_for('report'))

@app.route('/add_manual_record', methods=['GET', 'POST'])
@login_required
def add_manual_record():
    if current_user.role != 'admin':
        flash('Access denied: Admins only', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        date_str = request.form.get('date')
        punch_in = request.form.get('punch_in')
        punch_out = request.form.get('punch_out')
        
        # Validate required fields
        if not all([user_id, date_str, punch_in]):
            flash('Please fill in all required fields', 'danger')
            return redirect(request.url)
        
        try:
            # Parse and combine date with time
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            punch_in_time = datetime.strptime(f"{date_str} {punch_in}", "%Y-%m-%d %H:%M")
            
            # Handle punch out if provided
            punch_out_time = None
            hours_worked = None
            if punch_out:
                punch_out_time = datetime.strptime(f"{date_str} {punch_out}", "%Y-%m-%d %H:%M")
                if punch_out_time <= punch_in_time:
                    flash('Punch out time must be after punch in time', 'danger')
                    return redirect(request.url)
                
                # Calculate hours worked
                delta = punch_out_time - punch_in_time
                hours_worked = f"{delta.total_seconds() / 3600:.2f}"
            
            # Create new record
            new_record = PunchRecord(
                user_id=user_id,
                punch_in_time=punch_in_time,
                punch_out_time=punch_out_time,
                total_hours_worked=hours_worked,
                date=date_obj,
                is_manual=True
            )
            
            db.session.add(new_record)
            db.session.commit()
            
            flash('Manual record added successfully!', 'success')
            return redirect(url_for('report'))
            
        except ValueError as e:
            db.session.rollback()
            flash(f'Invalid date or time format: {str(e)}', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding manual record: {str(e)}', 'danger')
            app.logger.error(f"Error adding manual record: {str(e)}", exc_info=True)
    
    # For GET request, show the form
    users = User.query.order_by(User.username).all()
    return render_template('add_manual_record.html', users=users)

@app.route('/download_csv')
@login_required
def download_csv():
    if current_user.role != 'admin':
        flash('Access denied: Admins only', 'danger')
        return redirect(url_for('report'))

    record_ids = request.args.get('records')
    if not record_ids:
        flash('No records selected for export.', 'warning')
        return redirect(url_for('report'))

    record_ids = [int(id) for id in record_ids.split(',')]

    # Query selected records with all needed fields
    records = db.session.query(
        User.user_id, 
        User.username, 
        User.role,
        PunchRecord.punch_in_time, 
        PunchRecord.punch_out_time,
        PunchRecord.total_hours_worked
    ).join(PunchRecord).filter(PunchRecord.punch_id.in_(record_ids)).all()

    # Generate CSV data
    output = io.StringIO()
    csv_writer = csv.writer(output, delimiter=',')
    csv_writer.writerow(['ID', 'Name', 'Role', 'Punch In', 'Punch Out', 'Total Hours Worked'])
    
    for record in records:
        # Format times consistently
        punch_in = record.punch_in_time.strftime('%Y-%m-%d %H:%M') if record.punch_in_time else ''
        punch_out = record.punch_out_time.strftime('%Y-%m-%d %H:%M') if record.punch_out_time else ''
        
        # Convert time worked to decimal hours
        hours_worked = 0.0
        if record.total_hours_worked:
            try:
                # If it's already in decimal format
                hours_worked = float(record.total_hours_worked)
            except ValueError:
                # If it's in "M0/S2" or similar format
                if 'h' in record.total_hours_worked or 'm' in record.total_hours_worked or 's' in record.total_hours_worked:
                    # Parse time string like "1h 30m 15s"
                    parts = {'h': 0, 'm': 0, 's': 0}
                    for part in record.total_hours_worked.split():
                        if 'h' in part:
                            parts['h'] = float(part.replace('h', ''))
                        elif 'm' in part:
                            parts['m'] = float(part.replace('m', ''))
                        elif 's' in part:
                            parts['s'] = float(part.replace('s', ''))
                    hours_worked = parts['h'] + (parts['m'] / 60) + (parts['s'] / 3600)
                else:
                    # Fallback to 0 if format is unrecognized
                    hours_worked = 0.0
        
        csv_writer.writerow([
            record.user_id,
            record.username,
            record.role,
            punch_in,
            punch_out,
            f"{hours_worked:.2f}"  # Format as decimal with 2 places
        ])

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_report.csv"}
    )



@app.route('/clear_time_card/<int:punch_id>', methods=['POST'])
@login_required
def clear_time_card(punch_id):
    if current_user.role != 'admin':
        flash('Access denied: Admins only', 'danger')
        return redirect(url_for('dashboard'))

    try:
        # Delete only the specific punch record
        PunchRecord.query.filter_by(punch_id=punch_id).delete()
        db.session.commit()
        flash('Time entry deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting time entry: {e}", 'danger')

    return redirect(url_for('report'))

# ----- Authentication Routes -----

# Register route
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')  # Ensure this matches the form field name
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role', 'employee')  # Default to 'employee'

        # Check if the username already exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username already exists', 'danger')
            return redirect(url_for('register'))

        # Check if the email already exists
        existing_email = User.query.filter_by(email=email).first()
        if existing_email:
            flash('Email already exists', 'danger')
            return redirect(url_for('register'))

        # Create a new user
        new_user = User(username=username, email=email, role=role)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()

        flash('Registration successful', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')




# Login route
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()

        if user and user.check_password(password):  # This should work now
            login_user(user)
            flash('Login successful', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password', 'danger')

    return render_template('login.html')
# Logout route
@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    logout_user()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))
# ----- Punch In/Out Routes -----

@app.route('/set_timezone', methods=['POST'])
@login_required
def set_timezone():
    offset = request.form.get('timezone_offset', type=int)
    if offset is not None:
        current_user.timezone_offset = offset
        db.session.commit()
    return '', 204

from datetime import datetime, timezone, timedelta
import pytz

@app.route('/punch_with_location')
@login_required
def punch_with_location_page():
    """Display the punch with location form"""
    is_punched_in = PunchRecord.query.filter_by(
        user_id=current_user.user_id, 
        punch_out_time=None
    ).first() is not None
    
    return render_template(
        'punch_with_location.html', 
        is_punched_in=is_punched_in,
        google_maps_api_key=GOOGLE_MAPS_API_KEY
    )

@app.route('/punch_with_location', methods=['POST'])
@login_required
def punch_with_location():
    """Handle the punch with location form submission"""
    # Check if already punched in
    if PunchRecord.query.filter_by(user_id=current_user.user_id, punch_out_time=None).first():
        flash('You are already punched in', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get form data
    project_id = request.form.get('project_id')
    latitude = request.form.get('latitude')
    longitude = request.form.get('longitude')
    location_address = request.form.get('location_address')
    
    # Handle file upload
    progress_photo = None
    if 'progress_photo' in request.files:
        file = request.files['progress_photo']
        if file and file.filename != '' and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            unique_filename = f"{current_user.user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(file_path)
            progress_photo = unique_filename
    
    # Get project name for backward compatibility
    project_name = None
    if project_id:
        project = Project.query.get(project_id)
        if project:
            project_name = project.project_name
    
    # Create new punch record
    new_record = PunchRecord(
        user_id=current_user.user_id,
        punch_in_time=datetime.now(timezone.utc),
        date=datetime.now(timezone.utc).date(),
        latitude=latitude,
        longitude=longitude,
        location_address=location_address,
        project_id=project_id,
        project_name=project_name,  # Keep for backward compatibility
        progress_photo=progress_photo
    )
    
    db.session.add(new_record)
    db.session.commit()
    flash('Project Started successfully', 'success')
    return redirect(url_for('dashboard'))

@app.route('/punch_out_with_location', methods=['GET', 'POST'])
@login_required
def punch_out_with_location():
    if request.method == 'GET':
        return render_template('punch_out_with_location.html', 
                             google_maps_api_key=GOOGLE_MAPS_API_KEY)
    
    try:
        print("DEBUG: Punch out route called via POST")
        
        active_punch = PunchRecord.query.filter_by(
            user_id=current_user.user_id,
            punch_out_time=None
        ).first()
        
        if not active_punch:
            flash('No active punch in record found', 'danger')
            return redirect(url_for('dashboard'))
        
        print(f"DEBUG: Found active punch ID: {active_punch.punch_id}")
        
        # Check if form data is being received
        print(f"DEBUG: Form data: {request.form}")
        print(f"DEBUG: Files received: {request.files}")
        
        # Get location data from form
        latitude = request.form.get('latitude')
        longitude = request.form.get('longitude')
        location_address = request.form.get('location_address')
        
        print(f"DEBUG: Location data - lat: {latitude}, long: {longitude}, address: {location_address}")
        
        # Get project data from form
        task_description = request.form.get('task_description')
        progress_notes = request.form.get('progress_notes')
        
        print(f"DEBUG: Task: {task_description}, Notes: {progress_notes}")
        
        # Handle file upload for punch out - EXTENSIVE DEBUGGING
        progress_photo_out = None
        if 'progress_photo_out' in request.files:
            file = request.files['progress_photo_out']
            print(f"DEBUG: File object found: {file}")
            print(f"DEBUG: File filename: {file.filename}")
            print(f"DEBUG: File content type: {file.content_type}")
            print(f"DEBUG: File content length: {file.content_length}")
            
            if file and file.filename != '':
                print("DEBUG: File is not empty")
                if allowed_file(file.filename):
                    print("DEBUG: File type is allowed")
                    filename = secure_filename(file.filename)
                    unique_filename = f"punch_out_{current_user.user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                    
                    print(f"DEBUG: Attempting to save to: {file_path}")
                    
                    # Create uploads directory if it doesn't exist
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    
                    # Check directory permissions
                    upload_dir = app.config['UPLOAD_FOLDER']
                    print(f"DEBUG: Upload directory: {upload_dir}")
                    print(f"DEBUG: Directory exists: {os.path.exists(upload_dir)}")
                    print(f"DEBUG: Directory writable: {os.access(upload_dir, os.W_OK)}")
                    
                    try:
                        file.save(file_path)
                        print(f"DEBUG: File save attempted")
                        
                        # Verify file was actually saved
                        if os.path.exists(file_path):
                            file_size = os.path.getsize(file_path)
                            print(f"DEBUG: File successfully saved! Size: {file_size} bytes")
                            progress_photo_out = unique_filename
                        else:
                            print("DEBUG: ERROR: File save failed - file doesn't exist after save")
                    except Exception as save_error:
                        print(f"DEBUG: ERROR during file save: {str(save_error)}")
                else:
                    print(f"DEBUG: File type not allowed: {file.filename}")
            else:
                print("DEBUG: No file or empty filename received")
        else:
            print("DEBUG: No 'progress_photo_out' in request.files")
        
        print(f"DEBUG: Final progress_photo_out value: {progress_photo_out}")
        
        # Update punch record
        active_punch.punch_out_time = datetime.utcnow().replace(tzinfo=None)
        active_punch.latitude_out = latitude
        active_punch.longitude_out = longitude
        active_punch.location_address_out = location_address
        active_punch.task_description = task_description
        active_punch.progress_notes = progress_notes
        active_punch.progress_photo_out = progress_photo_out  # CRITICAL
        
        print(f"DEBUG: Setting progress_photo_out to: {progress_photo_out}")
        
        # Calculate hours worked
        if active_punch.punch_in_time and active_punch.punch_out_time:
            punch_in_naive = active_punch.punch_in_time.replace(tzinfo=None) if active_punch.punch_in_time.tzinfo else active_punch.punch_in_time
            punch_out_naive = active_punch.punch_out_time.replace(tzinfo=None) if active_punch.punch_out_time.tzinfo else active_punch.punch_out_time
            
            time_difference = punch_out_naive - punch_in_naive
            active_punch.total_hours_worked = round(time_difference.total_seconds() / 3600, 2)
        
        # Show what we're about to commit
        print(f"DEBUG: About to commit - progress_photo_out: {active_punch.progress_photo_out}")
        
        db.session.commit()
        print("DEBUG: Database commit successful")
        
        # Verify the data was actually saved
        updated_record = PunchRecord.query.get(active_punch.punch_id)
        print(f"DEBUG: AFTER COMMIT - progress_photo_out in DB: {updated_record.progress_photo_out}")
        
        flash('Project Stopped successfully with photo!', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error during punch out: {str(e)}")
        print(f"DEBUG: ERROR: {str(e)}")
        flash(f'Error during punch out: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    
@app.route('/debug_photos')
@login_required
def debug_photos():
    """Debug route to check photo data in database"""
    records = PunchRecord.query.filter_by(user_id=current_user.user_id).order_by(PunchRecord.punch_id.desc()).limit(5).all()
    
    debug_info = []
    for record in records:
        debug_info.append({
            'punch_id': record.punch_id,
            'punch_in_time': record.punch_in_time,
            'punch_out_time': record.punch_out_time,
            'progress_photo': record.progress_photo,
            'progress_photo_out': record.progress_photo_out,
            'has_punch_out_photo': bool(record.progress_photo_out)
        })
    
    return jsonify(debug_info)
    
@app.route('/debug_punch_data/<int:punch_id>')
@login_required
@admin_required
def debug_punch_data(punch_id):
    """Debug route to check punch data"""
    record = PunchRecord.query.get_or_404(punch_id)
    
    return jsonify({
        'punch_id': record.punch_id,
        'punch_in_time': record.punch_in_time.isoformat() if record.punch_in_time else None,
        'punch_out_time': record.punch_out_time.isoformat() if record.punch_out_time else None,
        'latitude': record.latitude,
        'longitude': record.longitude,
        'location_address': record.location_address,
        'latitude_out': record.latitude_out,
        'longitude_out': record.longitude_out,
        'location_address_out': record.location_address_out,
        'progress_photo': record.progress_photo,
        'progress_photo_out': record.progress_photo_out
    })
        
@app.route('/punch_in', methods=['POST'])
@login_required
def punch_in():
    if PunchRecord.query.filter_by(user_id=current_user.user_id, punch_out_time=None).first():
        flash('You are already punched in', 'warning')
        return redirect(url_for('dashboard'))

    # Use timezone.utc instead of UTC
    new_record = PunchRecord(
        user_id=current_user.user_id,
        punch_in_time=datetime.now(timezone.utc),  # ← Changed to timezone.utc
        date=datetime.now(timezone.utc).date()     # ← Changed to timezone.utc
    )
    
    db.session.add(new_record)
    db.session.commit()
    flash('Punched in successfully', 'success')
    return redirect(url_for('dashboard'))

@app.route('/punch_out', methods=['POST'])
@login_required
def punch_out():
    try:
        # Get the active punch record
        active_punch = PunchRecord.query.filter_by(
            user_id=current_user.user_id,
            punch_out_time=None
        ).first()
        
        if not active_punch:
            flash('No active punch in record found', 'danger')
            return redirect(url_for('dashboard'))
        
        # Get project data from form
        project_name = request.form.get('project_name')
        task_description = request.form.get('task_description')
        progress_notes = request.form.get('progress_notes')
        
        # Handle file upload
        progress_photo = None
        if 'progress_photo' in request.files:
            file = request.files['progress_photo']
            if file and file.filename != '':
                filename = secure_filename(file.filename)
                # Create unique filename
                unique_filename = f"{current_user.user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(file_path)
                progress_photo = unique_filename
        
        # Update punch record
        active_punch.punch_out_time = datetime.utcnow()
        active_punch.project_name = project_name
        active_punch.task_description = task_description
        active_punch.progress_notes = progress_notes
        active_punch.progress_photo = progress_photo
        
        # Calculate hours worked
        if active_punch.punch_in_time and active_punch.punch_out_time:
            time_difference = active_punch.punch_out_time - active_punch.punch_in_time
            active_punch.total_hours_worked = round(time_difference.total_seconds() / 3600, 2)
        
        db.session.commit()
        flash('Punched out successfully with project information', 'success')
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error during punch out: {str(e)}")
        flash(f'Error during punch out: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    
@app.route('/get_projects_json')
@login_required
def get_projects_json():
    """Return projects as JSON for dropdown"""
    try:
        projects = Project.query.filter_by(status='active').all()
        projects_list = [{'id': p.project_id, 'name': p.project_name} for p in projects]
        return jsonify(projects_list)
    except Exception as e:
        app.logger.error(f"Error getting projects: {str(e)}")
        return jsonify([])  # Return empty list on error
    
@app.route('/project_tracking')
@login_required
def project_tracking():
    # Get all punch records with project information
    records = db.session.query(
        PunchRecord.punch_id,
        PunchRecord.project_name,
        User.username,
        PunchRecord.date,
        PunchRecord.punch_in_time,
        PunchRecord.punch_out_time,
        PunchRecord.total_hours_worked,
        PunchRecord.task_description,
        PunchRecord.progress_photo
    ).join(User).filter(PunchRecord.project_name.isnot(None)).all()
    
    return render_template('project_tracking.html', records=records)

@app.route('/project_reports')
@login_required
def project_reports():
    # Get all punch records with project information
    records = db.session.query(
        PunchRecord.punch_id,
        PunchRecord.project_name,
        User.username,
        PunchRecord.date,
        PunchRecord.punch_out_time.isnot(None),  # Completed status
        PunchRecord.progress_photo
    ).join(User).filter(PunchRecord.project_name.isnot(None)).all()
    
    return render_template('project_reports.html', records=records)

@app.route('/project_details/<int:punch_id>')
@login_required
def project_details(punch_id):
    record = PunchRecord.query.get_or_404(punch_id)
    return render_template('project_details.html', record=record)
@app.route('/check_punch_status')
@login_required
def check_punch_status():
    record = PunchRecord.query.filter_by(
        user_id=current_user.user_id,
        punch_out_time=None
    ).first()
    return jsonify({'is_punched_in': record is not None})
# Protected Dashboard@app.route('/dashboard')
@app.route('/dashboard')
@login_required
def dashboard():
    # Initialize default values
    monthly_sales = 0.0
    sales_trends = []
    inventory_levels = {}
    current_record = None
    materials_by_project = {}
    project_progress = {}

    try:
        # Get current punch record
        current_record = PunchRecord.query.filter_by(
            user_id=current_user.user_id,
            punch_out_time=None
        ).first()

        if current_user.role == 'admin':
            # Connect to SQLite for materials data
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            # Create materials table if it doesn't exist
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS construction_materials (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    material_name TEXT NOT NULL,
                    category TEXT,
                    supplier TEXT,
                    unit_price REAL NOT NULL,
                    quantity_in_stock INTEGER NOT NULL,
                    minimum_stock_level INTEGER DEFAULT 0,
                    project_name TEXT,
                    date_added DATE DEFAULT CURRENT_DATE,
                    notes TEXT
                )
            """)
            
            # Get materials grouped by project
            cursor.execute("""
                SELECT project_name, material_name, quantity_in_stock, minimum_stock_level, unit_price
                FROM construction_materials 
                ORDER BY project_name, material_name
            """)
            materials = cursor.fetchall()
            
            # Organize materials by project
            for material in materials:
                project = material['project_name'] or 'General Stock'
                if project not in materials_by_project:
                    materials_by_project[project] = []
                materials_by_project[project].append(dict(material))
            
            # Calculate project progress (example - you'll need to define your own progress metrics)
            # For now, we'll use a simple example based on material availability
            for project, materials_list in materials_by_project.items():
                if project != 'General Stock':
                    total_materials = len(materials_list)
                    low_stock_count = sum(1 for m in materials_list 
                                        if m['quantity_in_stock'] <= m['minimum_stock_level'])
                    progress_percent = 100 - ((low_stock_count / total_materials) * 100) if total_materials > 0 else 0
                    project_progress[project] = min(max(progress_percent, 0), 100)
            
            conn.close()

    except Exception as e:
        app.logger.error(f"Error fetching dashboard data: {str(e)}")
        flash("Failed to fetch some dashboard data", "danger")

    return render_template(
        'dashboard.html',
        monthly_sales=monthly_sales,
        sales_trends=sales_trends,
        current_record=current_record,
        inventory_levels=inventory_levels,
        materials_by_project=materials_by_project,
        project_progress=project_progress,
        google_maps_api_key='AIzaSyC0fNpBUWXKaJJ32rQdgopgBTwwHaoVLX8',
    )
@app.route('/add_sample_materials')
@login_required
@admin_required
def add_sample_materials():
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Sample materials data
        sample_materials = [
            # Rosyln Project
            ('2x4 Lumber', 'Lumber', 'Home Depot', 5.99, 50, 20, 'Rosyln', 'Framing lumber'),
            ('Concrete Mix', 'Concrete', 'Lowe\'s', 8.49, 30, 15, 'Rosyln', 'Foundation work'),
            ('Drywall', 'Drywall', 'Building Supply', 12.99, 40, 25, 'Rosyln', 'Interior walls'),
            
            # Sabin Project
            ('Electrical Wire', 'Electrical', 'Electrical Supply', 0.89, 200, 100, 'Sabin', '12/2 Romex'),
            ('PVC Pipe', 'Plumbing', 'Plumbing World', 3.29, 60, 30, 'Sabin', '1/2" PVC'),
            ('Roof Shingles', 'Roofing', 'Roofing Supply', 45.99, 10, 5, 'Sabin', 'Architectural shingles'),
            
            # General Stock
            ('Nails', 'Hardware', 'Hardware Store', 0.05, 1000, 500, None, 'Various sizes'),
            ('Screws', 'Hardware', 'Hardware Store', 0.08, 800, 400, None, 'Wood and drywall'),
            ('Paint', 'Finishing', 'Paint Store', 25.99, 15, 10, None, 'White interior')
        ]
        
        cursor.executemany("""
            INSERT INTO construction_materials 
            (material_name, category, supplier, unit_price, quantity_in_stock, minimum_stock_level, project_name, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, sample_materials)
        
        conn.commit()
        conn.close()
        
        flash("Sample materials added successfully!", "success")
        
    except Exception as e:
        flash(f"Error adding sample materials: {str(e)}", "danger")
    
    return redirect(url_for('dashboard'))

@app.route('/edit_inventory/<int:product_id>', methods=['GET', 'POST'])
@login_required
def edit_inventory(product_id):
    print(f"\n=== EDIT INVENTORY REQUEST FOR ID: {product_id} ===")
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    try:
        # Debug: Show all inventory items first
        cursor.execute("SELECT * FROM inventory")
        all_items = cursor.fetchall()
        print("All inventory items before:", [dict(item) for item in all_items])
        
        if request.method == 'POST':
            print("POST request data:", request.form)
            
            # Get all form data
            brand_name = request.form['brand_name']
            product_name = request.form['product_name']
            price = float(request.form['price'])
            stock_quantity = int(request.form['stock_quantity'])
            date = request.form['date']
            
            # Execute update with explicit commit
            cursor.execute("""
                UPDATE inventory 
                SET brand_name = ?,
                    product_name = ?,
                    price = ?,
                    stock_quantity = ?,
                    date = ?,
                    last_updated = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (brand_name, product_name, price, stock_quantity, date, product_id))
            
            conn.commit()  # Explicit commit
            
            # Verify update
            cursor.execute("SELECT * FROM inventory WHERE id = ?", (product_id,))
            updated_item = cursor.fetchone()
            print("Updated item:", dict(updated_item) if updated_item else "None")
            
            flash("Inventory item updated successfully!", "success")
            return redirect(url_for('inventory'))
        
        # GET request handling
        cursor.execute("SELECT * FROM inventory WHERE id = ?", (product_id,))
        item = cursor.fetchone()
        print("Found item:", dict(item) if item else "None")
        
        if not item:
            flash("Inventory item not found", "danger")
            return redirect(url_for('inventory'))
            
        # Prepare brand data for template
        brands = ['Ganadara', 'Sushi Sama & POKE', 'VUA', 'ONIGIRI', 'Shawarmaz', 'Other']
        other_brand = item['brand_name'] if item['brand_name'] not in brands else ''
        
        return render_template('edit_inventory.html', 
                            product=item,
                            brands=brands,
                            other_brand=other_brand)

    except Exception as e:
        conn.rollback()  # Rollback on error
        print(f"Error in edit_inventory: {str(e)}")
        flash(f"An error occurred: {str(e)}", "danger")
        return redirect(url_for('inventory'))
    finally:
        # Verify changes after operation
        cursor.execute("SELECT * FROM inventory")
        print("All inventory items after:", [dict(item) for item in cursor.fetchall()])
        conn.close()

# Add this temporary route to get a valid ID
@app.route('/test_edit')
def test_edit():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM inventory LIMIT 1")
    item_id = cursor.fetchone()[0]
    conn.close()
    return redirect(url_for('edit_inventory', product_id=item_id))

@app.route('/edit_grocery/<int:item_id>', methods=['GET', 'POST'])
@login_required
def edit_grocery(item_id):
    if request.method == 'POST':
        try:
            brand_name = request.form['brand_name']
            product_name = request.form['product_name']
            price = float(request.form['price'])
            stock_quantity = int(request.form['stock_quantity'])
            date = request.form['date']

            with sqlite3.connect(db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE groceries 
                    SET brand_name=?, product_name=?, price=?, stock_quantity=?, date=?
                    WHERE id=?
                """, (brand_name, product_name, price, stock_quantity, date, item_id))
                conn.commit()
            
            flash("Grocery item updated successfully!", "success")
            return redirect(url_for('inventory'))
            
        except Exception as e:
            flash(f"Error updating item: {str(e)}", "danger")
    
    # GET request - show edit form
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM groceries WHERE id=?", (item_id,))
            item = cursor.fetchone()
            
        if not item:
            flash("Grocery item not found", "danger")
            return redirect(url_for('inventory'))
            
        return render_template('edit_grocery.html', item=item)
        
    except Exception as e:
        flash(f"Error retrieving item: {str(e)}", "danger")
        return redirect(url_for('inventory'))


    # Get only outstanding items (remaining quantities)
    placeholders = ','.join(['?'] * len(selected_order_ids))
    query = f"""
        SELECT 
            o.order_id,
            o.customer_name,
            o.customer_address,
            o.tags,
            oi.product_name,
            (oi.quantity - COALESCE(oi.fulfilled_qty, 0)) as remaining_quantity,
            oi.price
        FROM orders o
        JOIN order_items oi ON o.order_id = oi.order_id
        WHERE o.order_id IN ({placeholders})
          AND (oi.quantity - COALESCE(oi.fulfilled_qty, 0)) > 0
        ORDER BY o.customer_address, oi.product_name
    """
    cursor.execute(query, selected_order_ids)
    rows = cursor.fetchall()
    
    # Get inventory levels for all products in the picking list
    product_names = list(set([row[4] for row in rows]))  # Get unique product names
    inventory_levels = {}
    
    if product_names:
        inventory_placeholders = ','.join(['?'] * len(product_names))
        inventory_query = f"""
            SELECT product_name, SUM(stock_quantity) as total_stock 
            FROM inventory 
            WHERE product_name IN ({inventory_placeholders})
            GROUP BY product_name
        """
        cursor.execute(inventory_query, product_names)
        inventory_data = cursor.fetchall()
        inventory_levels = {item[0]: item[1] for item in inventory_data}
    
    conn.close()

    # Process data for template (your existing code)
    product_totals = defaultdict(int)
    address_groups = defaultdict(lambda: {
        'customer_name': '',
        'address': '',
        'tags': '',
        'products': []
    })

    for row in rows:
        order_id, customer_name, address, tags, product_name, remaining_quantity, price = row
        product_totals[product_name] += remaining_quantity
        
        if address not in address_groups:
            address_groups[address] = {
                'customer_name': customer_name,
                'address': address,
                'tags': tags,
                'products': []
            }
        address_groups[address]['products'].append({
            'name': product_name,
            'quantity': remaining_quantity,
            'price': price
        })

    return render_template('picking_list.html',
                         selected_order_count=len(selected_order_ids),
                         product_totals=product_totals,
                         address_groups=address_groups.values(),
                         inventory_levels=inventory_levels)  # Add inventory levels to template
    
@app.route('/materials')
@login_required
def materials():
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Check if table exists, create if not
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS construction_materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                material_name TEXT NOT NULL,
                category TEXT,
                supplier TEXT,
                unit_price REAL NOT NULL,
                quantity_in_stock INTEGER NOT NULL,
                minimum_stock_level INTEGER DEFAULT 0,
                project_name TEXT,
                date_added DATE DEFAULT CURRENT_DATE,
                notes TEXT
            )
        """)
        
        cursor.execute("SELECT * FROM construction_materials")
        materials = cursor.fetchall()
        
        conn.close()
        
        return render_template('materials.html', materials=materials)
        
    except Exception as e:
        flash(f"Error accessing materials: {str(e)}", "danger")
        return redirect(url_for('dashboard'))

@app.route('/add_material', methods=['POST'])
@login_required
def add_material():
    try:
        material_name = request.form['material_name']
        category = request.form['category']
        supplier = request.form['supplier']
        unit_price = float(request.form['unit_price'])
        quantity = int(request.form['quantity'])
        min_stock = int(request.form.get('min_stock', 0))
        project_name = request.form.get('project_name', '')
        notes = request.form.get('notes', '')
        
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO construction_materials 
            (material_name, category, supplier, unit_price, quantity_in_stock, minimum_stock_level, project_name, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (material_name, category, supplier, unit_price, quantity, min_stock, project_name, notes))
        
        conn.commit()
        conn.close()
        
        flash("Material added successfully!", "success")
        
    except Exception as e:
        flash(f"Error adding material: {str(e)}", "danger")
    
    return redirect(url_for('materials'))

@app.route('/inventory')
@login_required
def inventory():
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Fetch inventory items - explicitly name the id column as product_id
    cursor.execute("""
        SELECT id AS product_id, product_name, brand_name, price, stock_quantity, date 
        FROM inventory
    """)
    inventory_items = cursor.fetchall()

    # Fetch grocery items
    cursor.execute("""
        SELECT id, product_name, brand_name, price, stock_quantity, date 
        FROM groceries
    """)
    grocery_items = cursor.fetchall()

    conn.close()

    return render_template(
        'inventory.html', 
        inventory_items=inventory_items, 
        grocery_items=grocery_items
    )
@app.route('/delete_inventory/<int:product_id>', methods=['POST'])
@login_required
def delete_inventory(product_id):
    if current_user.role != 'admin':
        flash("Access denied: Admins only", "danger")
        return redirect(url_for('inventory'))

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM inventory WHERE id = ?", (product_id,))
        conn.commit()
        flash("Inventory item deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting inventory item: {e}", "danger")
    finally:
        if conn:
            conn.close()

    return redirect(url_for('inventory'))

@app.route('/export_inventory', methods=['POST'])
@login_required
def export_inventory():
    if current_user.role != 'admin':
        flash("Access denied: Admins only", "danger")
        return redirect(url_for('inventory'))

    export_type = request.form.get('export_type')
    selected_items = request.form.getlist('selected_items')

    if not selected_items:
        flash("No items selected for export", "warning")
        return redirect(url_for('inventory'))

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    try:
        placeholders = ','.join(['?'] * len(selected_items))
        
        if export_type == 'groceries':
            query = f"SELECT * FROM groceries WHERE id IN ({placeholders})"
            filename = "groceries_export.csv"
        else:
            query = f"SELECT * FROM inventory WHERE id IN ({placeholders})"
            filename = "inventory_export.csv"

        cursor.execute(query, selected_items)
        items = cursor.fetchall()
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Brand Name", "Product Name", "Price", "Stock Quantity", "Date"])
        
        for item in items:
            writer.writerow([
                item['id'],
                item['brand_name'],
                item['product_name'],
                item['price'],
                item['stock_quantity'],
                item['date']
            ])

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-type"] = "text/csv"
        return response

    except Exception as e:
        flash(f"Error during export: {str(e)}", "danger")
        return redirect(url_for('inventory'))
    finally:
        conn.close()

@app.route('/add_inventory_item', methods=['POST'])
@login_required
def add_inventory_item():
    if request.method == 'POST':
        try:
            # Get form data
            item_type = request.form['item_type']
            brand_name = request.form['brand_name']
            product_name = request.form['product_name']
            price = request.form['price']
            stock_quantity = request.form['stock_quantity']
            date = request.form['date']
            
            if brand_name == 'Other':
                brand_name = request.form.get('other_brand', 'Unknown Brand')

            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            
            if item_type == 'inventory':
                cursor.execute("""
                    INSERT INTO inventory 
                    (product_name, brand_name, price, quantity, stock_quantity, date)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    product_name,
                    brand_name,
                    float(price),
                    float(price),  # Using price as quantity since both are required
                    int(stock_quantity),
                    date
                ))
            else:
                cursor.execute("""
                    INSERT INTO groceries 
                    (product_name, brand_name, price, quantity, stock_quantity, date)
                    VALUES (?, ?, ?, ?, ?, ?)
                """, (
                    product_name,
                    brand_name,
                    float(price),
                    float(price),  # Using price as quantity since both are required
                    int(stock_quantity),
                    date
                ))
            
            conn.commit()
            flash("Item added successfully!", "success")
        except sqlite3.IntegrityError as e:
            flash(f"Error: {str(e)}", "danger")
        except Exception as e:
            flash(f"Error adding item: {str(e)}", "danger")
        finally:
            if 'conn' in locals():
                conn.close()
    
    return redirect(url_for('inventory'))

@app.route('/debug_db')
def debug_db():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    results = {}
    for table in tables:
        cursor.execute(f"SELECT * FROM {table[0]};")
        results[table[0]] = cursor.fetchall()
    conn.close()
    return jsonify(results)


# Add this temporary route to check your schema
@app.route('/debug_schema')
def debug_schema():
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(inventory)")
    inventory_columns = cursor.fetchall()
    cursor.execute("PRAGMA table_info(groceries)")
    grocery_columns = cursor.fetchall()
    conn.close()
    return jsonify({
        'inventory': inventory_columns,
        'groceries': grocery_columns
    })

@app.route('/add_grocery', methods=['GET', 'POST'])
@login_required
def add_grocery():
    if request.method == 'POST':
        product_name = request.form['product_name']
        brand_name = request.form['brand_name']
        price = request.form['price']
        stock_quantity = request.form['stock_quantity']
        date = request.form['date']

        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO groceries (product_name, brand_name, price, stock_quantity, date)
                VALUES (?, ?, ?, ?, ?)
            """, (product_name, brand_name, price, stock_quantity, date))
            conn.commit()
        except sqlite3.Error as e:
            flash(f"Error adding grocery item: {e}", "danger")
        finally:
            conn.close()

        flash("Grocery item added successfully!", "success")
        return redirect(url_for('inventory'))  

    return render_template('add_grocery.html')

@app.route('/delete_grocery/<int:item_id>', methods=['POST'])
@login_required
def delete_grocery(item_id):
    if current_user.role != 'admin':
        flash("Access denied: Admins only", "danger")
        return redirect(url_for('inventory'))

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM groceries WHERE id = ?", (item_id,))
        conn.commit()
        conn.close()
        flash("Grocery item deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting grocery item: {e}", "danger")

    return redirect(url_for('inventory'))




@app.route('/food_cost_calculator', methods=['GET', 'POST'])
@login_required
def food_cost_calculator():
    # Authentication check
    if current_user.role != 'admin':
        flash('Permission denied', 'danger')
        return redirect(url_for('dashboard'))

    conn = None
    try:
        conn = sqlite3.connect('instance/rrtables.db')
        cursor = conn.cursor()
        
        # Ensure all required columns exist
        ensure_columns_exist(cursor)
        
        # Handle POST requests
        if request.method == 'POST':
            # Handle item deletion
            if 'delete_item' in request.form:
                return handle_delete(cursor, conn)
            
            # Process and validate form data
            form_data = get_form_data(request)
            if not form_data:
                return redirect(url_for('food_cost_calculator'))
            
            # Perform all calculations
            calculations = calculate_all_values(form_data)
            
            # Handle different form actions
            if 'calculate' in request.form:
                return show_calculation_results(cursor, form_data, calculations)
            elif 'save_item' in request.form:
                return save_item(cursor, conn, form_data, calculations)
            elif 'update_item' in request.form:
                return update_item(cursor, conn, form_data, calculations)
        
        # GET request - show empty form with saved items
        return show_empty_form(cursor)

    except Exception as e:
        if conn:
            conn.rollback()
        flash(f"Error: {str(e)}", "danger")
        return redirect(url_for('food_cost_calculator'))
    finally:
        if conn:
            conn.close()

# ========== HELPER FUNCTIONS ==========

def ensure_columns_exist(cursor):
    """Ensure all required database columns exist"""
    cursor.execute("PRAGMA table_info(food_cost_items)")
    existing_columns = {col[1] for col in cursor.fetchall()}
    required_columns = {
        'revenue_after_fees': 'REAL',
        'total_costs': 'REAL',
        'profit': 'REAL',
        'profit_margin': 'REAL',
        'saved_quantity': 'INTEGER',
        'brand': 'TEXT',
        'unit_type': 'TEXT'
    }
    
    for col, col_type in required_columns.items():
        if col not in existing_columns:
            cursor.execute(f"ALTER TABLE food_cost_items ADD COLUMN {col} {col_type}")

def get_form_data(request):
    """Extract and validate form data"""
    try:
        form_data = {
            'product_name': request.form.get('product_name', '').strip(),
            'product_code': request.form.get('product_code', '').strip(),
            'price_per_unit': float(request.form.get('price_per_unit', 0)),
            'unit_type': request.form.get('unit_type', 'kg'),
            'raw_material_cost': float(request.form.get('raw_material_cost', 0)),
            'waste_value': float(request.form.get('waste_value', 0)),
            'waste_type': request.form.get('waste_type', 'dollar'),
            'labor_cost': float(request.form.get('labor_cost', 0)),
            'fixed_cost': float(request.form.get('fixed_cost', 0)),
            'fixed_cost_type': request.form.get('fixed_cost_type', 'dollar'),
            'kickback_percentage': float(request.form.get('kickback_percentage', 0)),
            'fee_applies': bool(int(request.form.get('fee_applies', 1))),
            'quantity': int(request.form.get('quantity', 1)),
            'brand': request.form.get('brand', '').strip(),
            'item_id': request.form.get('item_id', '')
        }
        
        # Validate required fields
        required_fields = ['product_name', 'product_code', 'price_per_unit',
                         'raw_material_cost', 'labor_cost', 'fixed_cost']
        for field in required_fields:
            if not form_data[field]:
                flash(f"{field.replace('_', ' ').title()} is required", "danger")
                return None
                
        return form_data
        
    except ValueError:
        flash("Invalid numeric values", "danger")
        return None

def calculate_all_values(form_data):
    """Perform all financial calculations"""
    # Waste calculation
    if form_data['waste_type'] == 'percentage':
        waste_cost = (form_data['waste_value'] / 100) * form_data['raw_material_cost']
        waste_display = f"{form_data['waste_value']}% of ${form_data['raw_material_cost']:.2f} = ${waste_cost:.2f}"
    else:
        waste_cost = form_data['waste_value']
        waste_display = f"${form_data['waste_value']:.2f}"

    # Fixed cost calculation
    if form_data['fixed_cost_type'] == 'percentage':
        fixed_cost = (form_data['fixed_cost'] / 100) * form_data['price_per_unit']
        fixed_cost_display = f"{form_data['fixed_cost']}% of ${form_data['price_per_unit']:.2f} = ${fixed_cost:.2f}"
    else:
        fixed_cost = form_data['fixed_cost']
        fixed_cost_display = f"${form_data['fixed_cost']:.2f}"

    # Total calculations
    total_cost = (form_data['raw_material_cost'] + waste_cost + 
                 form_data['labor_cost'] + fixed_cost)
    
    fee_amount = form_data['price_per_unit'] * 0.08 if form_data['fee_applies'] else 0
    revenue_after_fees = form_data['price_per_unit'] - fee_amount
    profit = revenue_after_fees - total_cost
    profit_margin = (profit / revenue_after_fees) * 100 if revenue_after_fees > 0 else 0

    return {
        'waste_cost': waste_cost,
        'waste_display': waste_display,
        'fixed_cost': fixed_cost,
        'fixed_cost_display': fixed_cost_display,
        'total_cost': total_cost,
        'fee_amount': fee_amount,
        'revenue_after_fees': revenue_after_fees,
        'profit': profit,
        'profit_margin': profit_margin
    }

def handle_delete(cursor, conn):
    """Handle item deletion"""
    item_id = request.form.get('item_id')
    if not item_id:
        flash("Item ID missing", "danger")
        return redirect(url_for('food_cost_calculator'))
    
    cursor.execute('DELETE FROM food_cost_items WHERE id = ? AND user_id = ?', 
                  (item_id, current_user.get_id()))
    conn.commit()
    flash('Item deleted', 'success')
    return redirect(url_for('food_cost_calculator'))

def show_calculation_results(cursor, form_data, calculations):
    """Display calculation results without saving"""
    saved_items, brands = fetch_saved_items(cursor)
    context = {
        **form_data,
        **calculations,
        'calculated': True,
        'saved_items': saved_items,
        'brands': brands
    }
    return render_template('food_cost_calculator.html', **context)

def save_item(cursor, conn, form_data, calculations):
    """Save new item to database"""
    db_data = {
        'user_id': current_user.get_id(),
        **form_data,
        **calculations
    }
    
    columns = ', '.join(db_data.keys())
    placeholders = ', '.join(['?'] * len(db_data))
    cursor.execute(f'INSERT INTO food_cost_items ({columns}) VALUES ({placeholders})', 
                  tuple(db_data.values()))
    conn.commit()
    flash('Item saved successfully', 'success')
    return redirect(url_for('food_cost_calculator'))

def update_item(cursor, conn, form_data, calculations):
    """Update existing item in database"""
    if not form_data.get('item_id'):
        flash("Item ID missing for update", "danger")
        return redirect(url_for('food_cost_calculator'))
    
    db_data = {
        **form_data,
        **calculations
    }
    
    set_clause = ', '.join([f"{key} = ?" for key in db_data if key != 'item_id'])
    values = [val for key, val in db_data.items() if key != 'item_id']
    values.extend([form_data['item_id'], current_user.get_id()])
    
    cursor.execute(f'UPDATE food_cost_items SET {set_clause} WHERE id = ? AND user_id = ?', 
                  tuple(values))
    conn.commit()
    flash('Item updated successfully', 'success')
    return redirect(url_for('food_cost_calculator'))

def fetch_saved_items(cursor):
    """Retrieve all saved items for the current user"""
    cursor.execute('''
        SELECT * FROM food_cost_items 
        WHERE user_id = ? 
        ORDER BY brand, product_name
    ''', (current_user.get_id(),))
    saved_items = cursor.fetchall()
    brands = sorted({item[22] for item in saved_items if item[22]})
    return saved_items, brands

def show_empty_form(cursor):
    """Display empty form with saved items list"""
    saved_items, brands = fetch_saved_items(cursor)
    return render_template('food_cost_calculator.html',
                         calculated=False,
                         saved_items=saved_items,
                         brands=brands)

@app.route('/load_food_cost_item/<int:item_id>')
@login_required
def load_food_cost_item(item_id):
    conn = sqlite3.connect('instance/rrtables.db')
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT id, product_name, product_code, price_per_kg, unit_type,
                   raw_material_cost, waste_value, waste_type, waste_cost,
                   labor_cost, fixed_cost, fixed_cost_type, kickback_percentage, fee_applies,
                   revenue_after_fees, total_costs, profit, profit_margin, saved_quantity, brand
            FROM food_cost_items
            WHERE id = ? AND user_id = ?
        ''', (item_id, current_user.get_id()))
        
        item = cursor.fetchone()
        
        if item:
            item_data = {
                'id': item[0],
                'product_name': item[1],
                'product_code': item[2],
                'price_per_unit': item[3],
                'unit_type': item[4],
                'raw_material_cost': item[5],
                'waste_value': item[6],
                'waste_type': item[7],
                'waste_cost': item[8],
                'labor_cost': item[9],
                'fixed_cost': item[10],
                'fixed_cost_type': item[11],
                'kickback_percentage': item[12] * 100,
                'fee_applies': item[13],
                'revenue_after_fees': item[14],
                'total_costs': item[15],
                'profit': item[16],
                'profit_margin': item[17],
                'saved_quantity': item[18],
                'brand': item[19] if len(item) > 19 else ''
            }
            return jsonify(item_data)
        else:
            return jsonify({'error': 'Item not found'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()
@app.route('/export_food_cost_items', methods=['POST'])
@login_required
def export_food_cost_items():
    if current_user.role != 'admin':
        return jsonify({'error': 'Permission denied'}), 403

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        flash('Excel export requires openpyxl package', 'danger')
        return redirect(url_for('food_cost_calculator'))

    selected_ids = request.form.getlist('selected_items')
    if not selected_ids:
        flash('No items selected', 'warning')
        return redirect(url_for('food_cost_calculator'))

    try:
        conn = sqlite3.connect('instance/rrtables.db')
        cursor = conn.cursor()
        
        # Use parameterized query safely
        query = """
            SELECT 
                brand, product_name, product_code, price_per_kg, unit_type,
                raw_material_cost, waste_cost, labor_cost, fixed_cost,
                fee_applies, total_costs, profit_margin, saved_quantity
            FROM food_cost_items 
            WHERE id IN ({}) AND user_id = ?
            ORDER BY brand, product_name
        """.format(','.join(['?']*len(selected_ids)))
        
        cursor.execute(query, (*selected_ids, current_user.get_id()))
        items = cursor.fetchall()

        if not items:
            flash('No items found', 'warning')
            return redirect(url_for('food_cost_calculator'))

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Food Cost Analysis"

        # Styles
        header_style = {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill("solid", fgColor="4F81BD"),
            'border': Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            ),
            'alignment': Alignment(horizontal='center')
        }

        # Headers
        headers = [
            "Brand", "Name", "Code", "Price/Unit", 
            "Raw Materials", "Waste", "Labor", "Fixed Cost",
            "8% Fee", "Total Cost", "Margin", "Quantity"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # Data rows
        for row, item in enumerate(items, 2):
            quantity = item[12] or 1
            ws.cell(row=row, column=1, value=item[0])  # Brand
            ws.cell(row=row, column=2, value=item[1])  # Name
            ws.cell(row=row, column=3, value=item[2])  # Code
            ws.cell(row=row, column=4, value=f"${item[3]:.2f}/{item[4]}")  # Price
            ws.cell(row=row, column=5, value=item[5]).number_format = '$#,##0.00'
            ws.cell(row=row, column=6, value=item[6]).number_format = '$#,##0.00'
            ws.cell(row=row, column=7, value=item[7]).number_format = '$#,##0.00'
            ws.cell(row=row, column=8, value=item[8]).number_format = '$#,##0.00'
            ws.cell(row=row, column=9, value="Yes" if item[9] else "No")
            ws.cell(row=row, column=10, value=item[10]).number_format = '$#,##0.00'
            ws.cell(row=row, column=11, value=item[11]).number_format = '0.00%'
            ws.cell(row=row, column=12, value=quantity)

        # Auto-size columns
        for column in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in column
            ) + 2
            ws.column_dimensions[column[0].column_letter].width = max_length

        # Return file
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return Response(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=food_costs.xlsx'}
        )

    except Exception as e:
        flash(f'Export failed: {str(e)}', 'danger')
        return redirect(url_for('food_cost_calculator'))
    finally:
        conn.close()

@app.route('/production_tracking')
@login_required
@admin_required
def production_tracking():
    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id')
    
    # Base query with joinedload for user relationship
    query = db.session.query(ProductionRecord).join(User).options(
        db.joinedload(ProductionRecord.user))
    
    # Apply date filters if provided
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(ProductionRecord.production_date >= start_date_obj)
        except ValueError:
            flash('Invalid start date format', 'danger')
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(ProductionRecord.production_date <= end_date_obj)
        except ValueError:
            flash('Invalid end date format', 'danger')
    
    # Apply employee filter if provided
    if employee_id and employee_id.isdigit():
        query = query.filter(ProductionRecord.user_id == int(employee_id))
    
    # Get all users for filter dropdown
    all_users = User.query.order_by(User.username).all()
    
    # Execute query and group results
    records = query.order_by(
        User.username,
        ProductionRecord.production_date.desc(),
        ProductionRecord.recorded_at.desc()
    ).all()
    
    # Group records by employee
    grouped_records = {}
    for record in records:
        if record.user_id not in grouped_records:
            grouped_records[record.user_id] = []
        grouped_records[record.user_id].append(record)
    
    return render_template(
        'production_tracking.html',
        grouped_records=grouped_records,
        all_users=all_users,
        start_date=start_date,
        end_date=end_date,
        employee_id=employee_id
    )
@app.route('/delete_production_record/<int:record_id>', methods=['POST'])
@login_required
@admin_required
def delete_production_record(record_id):
    record = ProductionRecord.query.get_or_404(record_id)
    try:
        db.session.delete(record)
        db.session.commit()
        flash('Production record deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting record: {str(e)}', 'danger')
    return redirect(url_for('production_tracking'))

@app.route('/edit_production_record/<int:record_id>', methods=['POST'])
@login_required
@admin_required
def edit_production_record(record_id):
    record = ProductionRecord.query.get_or_404(record_id)
    try:
        record.product_name = request.form['product_name']
        if record.product_name != 'General Duties':
            record.quantity = float(request.form['quantity'])
            record.lot_number = request.form['lot_number']
        else:
            record.quantity = 0
            record.lot_number = 'N/A'
        db.session.commit()
        flash('Record updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating record: {str(e)}', 'danger')
    return redirect(url_for('production_tracking'))
@app.route('/get_production_record/<int:record_id>')
@login_required
@admin_required
def get_production_record(record_id):
    record = ProductionRecord.query.get_or_404(record_id)
    return jsonify({
        'product_name': record.product_name,
        'quantity': record.quantity,
        'lot_number': record.lot_number
    })




@app.route('/product_profitability')
@admin_required
def product_profitability():
    products = Product.query.all()
    
    cost_data = []
    for product in products:
        selling_price = product.selling_price or 0
        direct_costs = product.direct_costs_total or 0
        indirect_costs = product.indirect_costs or 0
        
        # Calculate profit if not set in DB
        profit = product.profit or (selling_price - direct_costs - indirect_costs)
        
        # Calculate margin (avoid division by zero)
        margin = 0
        if selling_price > 0:
            margin = (profit / selling_price) * 100
            
        cost_data.append({
            'product_id': product.id,
            'product_name': product.name,
            'selling_price': selling_price,
            'direct_costs': direct_costs,
            'indirect_costs': indirect_costs,
            'profit': profit,
            'margin': margin
        })
    
    return render_template('product_profitability.html', cost_data=cost_data)


@app.route('/purchase_orders')
@admin_required
def purchase_orders():
    status = request.args.get('status', 'all')
    page = request.args.get('page', 1, type=int)
    
    query = PurchaseOrder.query
    if status != 'all':
        query = query.filter_by(status=status)
    
    pos = query.order_by(PurchaseOrder.order_date.desc()).paginate(page=page, per_page=20)
    return render_template('purchase_orders.html', pos=pos, status=status)

@app.route('/create_po', methods=['GET', 'POST'])
@admin_required
def create_po():
    if request.method == 'POST':
        try:
            # Check if we're adding a new distributor
            if request.form.get('new_distributor_name'):
                distributor_name = request.form.get('new_distributor_name')
                contact_name = request.form.get('new_distributor_contact', '')
                
                if not distributor_name:
                    flash('Distributor name is required', 'error')
                    return redirect(url_for('create_po'))
                
                new_distributor = Distributor(
                    name=distributor_name,
                    contact_name=contact_name,
                    active=True
                )
                db.session.add(new_distributor)
                db.session.commit()
                distributor_id = new_distributor.id
            else:
                distributor_id = request.form.get('distributor_id')
                if not distributor_id:
                    flash('Please select or create a distributor', 'error')
                    return redirect(url_for('create_po'))
            
            # Get order date from form or use today's date
            order_date_str = request.form.get('order_date')
            order_date = datetime.strptime(order_date_str, '%Y-%m-%d').date() if order_date_str else datetime.now().date()
            
            expected_delivery_str = request.form.get('expected_delivery_date')
            notes = request.form.get('notes')
            
            # Convert expected delivery date string to date object
            expected_delivery_date = None
            if expected_delivery_str:
                expected_delivery_date = datetime.strptime(expected_delivery_str, '%Y-%m-%d').date()
            
            # Generate PO number
            last_po = PurchaseOrder.query.order_by(PurchaseOrder.id.desc()).first()
            po_number = f"PO-{datetime.now().year}-{(last_po.id + 1) if last_po else 1:03d}"
            
            new_po = PurchaseOrder(
                po_number=po_number,
                distributor_id=distributor_id,
                order_date=order_date,
                expected_delivery_date=expected_delivery_date,
                status='draft',
                notes=notes
            )
            db.session.add(new_po)
            db.session.commit()
            
            flash('Purchase order created successfully!', 'success')
            return redirect(url_for('edit_po', po_id=new_po.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating PO: {str(e)}', 'error')
            return redirect(url_for('create_po'))
    
    distributors = Distributor.query.filter_by(active=True).all()
    return render_template('create_po.html', distributors=distributors)

@app.route('/edit_po/<int:po_id>', methods=['GET', 'POST'])
@admin_required
def edit_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    if request.method == 'POST':
        if 'add_item' in request.form:
            try:
                item_type = request.form.get('item_type')
                quantity = float(request.form.get('quantity'))
                unit_price = float(request.form.get('unit_price'))
                unit_of_measure = request.form.get('unit_of_measure', 'unit')

                new_item = POItem(
                    po_id=po.id,
                    quantity=quantity,
                    unit_price=unit_price,
                    unit_of_measure=unit_of_measure
                )

                if item_type == 'product':
                    new_item.product_id = request.form.get('item_id')
                elif item_type == 'ingredient':
                    new_item.ingredient_id = request.form.get('item_id')
                else:  # manual entry
                    new_item.description = request.form.get('item_name')

                db.session.add(new_item)
                db.session.commit()
                flash('Item added to PO', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding item: {str(e)}', 'error')
            
            return redirect(url_for('edit_po', po_id=po.id))
        
        elif 'update_status' in request.form:
            try:
                new_status = request.form.get('status')
                if new_status in ['draft', 'open', 'pending', 'paid']:
                    po.status = new_status
                    db.session.commit()
                    flash('PO status updated', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating status: {str(e)}', 'error')
            return redirect(url_for('edit_po', po_id=po.id))
        
        elif 'update_order_date' in request.form:
            try:
                new_date_str = request.form.get('order_date')
                if new_date_str:
                    po.order_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
                    db.session.commit()
                    flash('Order date updated', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error updating order date: {str(e)}', 'error')
            return redirect(url_for('edit_po', po_id=po.id))
        
        elif 'submit_po' in request.form:
            try:
                po.status = 'open'
                db.session.commit()
                flash('Purchase order submitted successfully!', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error submitting PO: {str(e)}', 'error')
            return redirect(url_for('edit_po', po_id=po.id))
    
    products = Product.query.all()
    ingredients = Ingredient.query.all()
    
    return render_template('edit_po.html', 
                         po=po, 
                         products=products, 
                         ingredients=ingredients,
                         statuses=['draft', 'open', 'pending', 'paid'])

@app.route('/export_po/<int:po_id>/<format>')
@admin_required
def export_po(po_id, format):
    po = PurchaseOrder.query.get_or_404(po_id)
    hide_prices = request.args.get('hide_prices', 'false').lower() == 'true'
    
    def format_currency(value):
        return "${:,.2f}".format(float(value)) if not hide_prices else "---"
    
    if format == 'pdf':
        buffer = io.BytesIO()
        doc = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Set margins
        left_margin = 50
        right_margin = 50
        top_margin = 50
        bottom_margin = 50
        
        # Set up styles
        doc.setFont("Helvetica-Bold", 16)
        
        # Draw PO header
        doc.drawString(left_margin, height - top_margin - 20, f"PURCHASE ORDER: {po.po_number}")
        if hide_prices:
            doc.drawString(width - 200, height - top_margin - 20, "PRICES HIDDEN")
        
        doc.setFont("Helvetica", 12)
        y_position = height - top_margin - 50
        doc.drawString(left_margin, y_position, f"Distributor: {po.distributor.name}")
        y_position -= 20
        doc.drawString(left_margin, y_position, f"Order Date: {po.order_date.strftime('%Y-%m-%d')}")
        y_position -= 20
        doc.drawString(left_margin, y_position, f"Status: {po.status.upper()}")
        
        # Draw line items header
        y_position -= 40
        doc.setFont("Helvetica-Bold", 14)
        doc.drawString(left_margin, y_position, "ITEMS ORDERED")
        y_position -= 20
        
        # Table header
        doc.setFont("Helvetica-Bold", 10)
        doc.drawString(left_margin, y_position, "Item")
        doc.drawString(width - right_margin - 300, y_position, "Qty")
        if not hide_prices:
            doc.drawString(width - right_margin - 200, y_position, "Unit Price")
        doc.drawString(width - right_margin - 100, y_position, "Total")
        y_position -= 20
        
        # Draw items
        doc.setFont("Helvetica", 10)
        for item in po.items:
            if y_position < bottom_margin + 40:  # Leave room for total
                doc.showPage()
                y_position = height - top_margin - 20
                doc.setFont("Helvetica", 10)
            
            if item.product_id:
                name = item.product.name
            elif item.ingredient_id:
                name = item.ingredient.name
            else:
                name = item.description
                
            doc.drawString(left_margin, y_position, f"{name} ({item.unit_of_measure})")
            doc.drawString(width - right_margin - 300, y_position, str(item.quantity))
            if not hide_prices:
                doc.drawString(width - right_margin - 200, y_position, format_currency(item.unit_price))
            doc.drawString(width - right_margin - 100, y_position, format_currency(item.line_total))
            y_position -= 20
        
        # Draw total
        doc.setFont("Helvetica-Bold", 12)
        total = sum(item.line_total for item in po.items)
        doc.drawString(width - right_margin - 100, y_position, f"TOTAL: {format_currency(total)}")
        
        doc.save()
        buffer.seek(0)
        return Response(
            buffer,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment;filename=PO-{po.po_number}{'-no-prices' if hide_prices else ''}.pdf"}
        )
    elif format == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'PO Number', 'Distributor', 'Order Date', 'Status',
            'Item', 'Quantity', 'Unit', 'Unit Price', 'Line Total'
        ])
        
        # Write PO data and items
        for item in po.items:
            if item.product_id:
                name = item.product.name
            elif item.ingredient_id:
                name = item.ingredient.name
            else:
                name = item.description
                
            writer.writerow([
                po.po_number,
                po.distributor.name,
                po.order_date.strftime('%Y-%m-%d'),
                po.status,
                name,
                item.quantity,
                item.unit_of_measure,
                item.unit_price,
                item.line_total
            ])
        
        output.seek(0)
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename=PO-{po.po_number}.csv"}
        )
@app.route('/delete_po/<int:po_id>', methods=['POST'])
@admin_required
def delete_po(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    try:
        # Delete all items first to maintain referential integrity
        POItem.query.filter_by(po_id=po_id).delete()
        db.session.delete(po)
        db.session.commit()
        flash('Purchase order deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting PO: {str(e)}', 'error')
    return redirect(url_for('purchase_orders'))

@app.route('/delete_po_item/<int:item_id>', methods=['POST'])
@admin_required
def delete_po_item(item_id):
    item = POItem.query.get_or_404(item_id)
    po_id = item.po_id
    try:
        db.session.delete(item)
        db.session.commit()
        flash('Item removed from PO', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting item: {str(e)}', 'error')
    return redirect(url_for('edit_po', po_id=po_id))

@app.route('/export_po/<int:po_id>/csv')
@admin_required
def export_po_csv(po_id):
    po = PurchaseOrder.query.get_or_404(po_id)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'PO Number', 'Distributor', 'Order Date', 'Status',
        'Item', 'Quantity', 'Unit', 'Unit Price', 'Line Total'
    ])
    
    # Write PO data and items
    for item in po.items:
        if item.product_id:
            name = item.product.name
        elif item.ingredient_id:
            name = item.ingredient.name
        else:
            name = item.description
            
        writer.writerow([
            po.po_number,
            po.distributor.name,
            po.order_date.strftime('%Y-%m-%d'),
            po.status,
            name,
            item.quantity,
            item.unit_of_measure,
            item.unit_price,
            item.line_total
        ])
    
    output.seek(0)
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=PO-{po.po_number}.csv"}
    )

@app.route('/receive_po_item/<int:item_id>', methods=['POST'])
@admin_required
def receive_po_item(item_id):
    item = POItem.query.get_or_404(item_id)
    received_qty = float(request.form.get('received_qty'))
    
    item.received_quantity += received_qty
    
    if item.received_quantity >= item.quantity:
        item.status = 'fully_received'
    elif item.received_quantity > 0:
        item.status = 'partially_received'
    
    db.session.commit()
    flash('Received quantity updated', 'success')
    return redirect(url_for('edit_po', po_id=item.po_id))


@app.route('/production_calendar', methods=['GET', 'POST'])
@login_required
def production_calendar():
    if request.method == 'POST':
        try:
            # Get form data
            project_name = request.form['project_name']
            date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
            start_time = datetime.strptime(request.form['start_time'], '%H:%M').time() if request.form.get('start_time') else None
            end_time = datetime.strptime(request.form['end_time'], '%H:%M').time() if request.form.get('end_time') else None
            team_members = request.form.get('team_members', '')
            description = request.form.get('description', '')
            
            # Create new calendar event
            new_event = ProjectCalendar(
                project_name=project_name,
                date=date,
                start_time=start_time,
                end_time=end_time,
                team_members=team_members,
                description=description,
                created_by=current_user.user_id
            )
            
            db.session.add(new_event)
            db.session.commit()
            flash('Project event scheduled successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error scheduling project event: {str(e)}', 'danger')
            app.logger.error(f"Project scheduling error: {str(e)}")
        
        return redirect(url_for('production_calendar'))
    
    # Get existing events
    events = ProjectCalendar.query.order_by(
        ProjectCalendar.date,
        ProjectCalendar.start_time
    ).all()
    
    # Get all users for team member selection
    all_users = User.query.order_by(User.username).all()
    
    return render_template('production_calendar.html', 
                         events=events, 
                         all_users=all_users)

@app.route('/production_calendar/request_materials/<int:event_id>', methods=['POST'])
@login_required
def request_materials(event_id):
    try:
        # Get JSON data from request
        data = request.get_json()
        
        # Update the event with materials request
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Store materials request as JSON
        materials_data = {
            'material_name': data.get('material_name'),
            'quantity': data.get('quantity'),
            'unit': data.get('unit'),
            'priority': data.get('priority'),
            'notes': data.get('notes'),
            'requested_by': current_user.username,
            'requested_at': datetime.now().isoformat()
        }
        
        cursor.execute("""
            UPDATE project_calendar 
            SET materials_needed = ?, status = 'pending-materials'
            WHERE id = ?
        """, (json.dumps(materials_data), event_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Materials request submitted successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/production_calendar/delete/<int:event_id>', methods=['POST'])
@login_required
def delete_calendar_event(event_id):
    event = ProjectCalendar.query.get_or_404(event_id)
    
    # Only allow deletion by admin or creator
    if current_user.role != 'admin' and event.created_by != current_user.user_id:
        flash('You can only delete events you created', 'danger')
        return redirect(url_for('production_calendar'))
    
    db.session.delete(event)
    db.session.commit()
    flash('Calendar event deleted successfully', 'success')
    return redirect(url_for('production_calendar'))

@app.route('/production_calendar/ical')
@login_required
def production_calendar_ical():
    events = ProjectCalendar.query.order_by(ProjectCalendar.date).all()
    
    ical = "BEGIN:VCALENDAR\nVERSION:2.0\nPRODID:-//Construction//Project Calendar//EN\n"
    
    for event in events:
        ical += "BEGIN:VEVENT\n"
        ical += f"UID:{event.id}@construction.com\n"
        ical += f"DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}\n"
        ical += f"DTSTART;VALUE=DATE:{event.date.strftime('%Y%m%d')}\n"
        
        if event.start_time and event.end_time:
            start_dt = datetime.combine(event.date, event.start_time)
            end_dt = datetime.combine(event.date, event.end_time)
            ical += f"DTSTART:{start_dt.strftime('%Y%m%dT%H%M%S')}\n"
            ical += f"DTEND:{end_dt.strftime('%Y%m%dT%H%M%S')}\n"
        
        ical += f"SUMMARY:Project - {event.project_name}\n"
        ical += f"DESCRIPTION:Team: {event.team_members or 'Not assigned'}\\n"
        ical += f"Status: {event.status}\\n"
        ical += f"Notes: {event.description or 'No notes'}\n"
        ical += "END:VEVENT\n"
    
    ical += "END:VCALENDAR"
    
    response = make_response(ical)
    response.headers['Content-Type'] = 'text/calendar'
    response.headers['Content-Disposition'] = 'attachment; filename=project_calendar.ics'
    return response

@app.route('/production_calendar/materials_form/<int:event_id>')
@login_required
def materials_request_form(event_id):
    # Get the event using SQLite since you're using direct SQLite queries
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT * FROM project_calendar WHERE id = ?", (event_id,))
    event = cursor.fetchone()
    
    conn.close()
    
    if not event:
        flash('Event not found', 'danger')
        return redirect(url_for('production_calendar'))
    
    return render_template('materials_request_form.html', event=event)


@app.route('/production_calendar/view')
@login_required
def production_calendar_view():
    # Get events for the current month
    today = datetime.now()
    first_day = today.replace(day=1)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT * FROM project_calendar 
        WHERE date BETWEEN ? AND ?
        ORDER BY date, start_time
    """, (first_day.strftime('%Y-%m-%d'), last_day.strftime('%Y-%m-%d')))
    
    events = cursor.fetchall()
    conn.close()
    
    return render_template('calendar_view.html', 
                         events=events, 
                         current_month=today.strftime('%B %Y'),
                         calendar=generate_calendar(today.year, today.month, events))

def generate_calendar(year, month, events):
    # Create a calendar matrix
    cal = calendar.monthcalendar(year, month)
    calendar_data = []
    
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'day': None, 'events': []})
            else:
                day_date = f"{year}-{month:02d}-{day:02d}"
                day_events = [event for event in events if event['date'] == day_date]
                week_data.append({'day': day, 'events': day_events})
        calendar_data.append(week_data)
    
    return calendar_data
@app.route('/production_calendar/approve_materials/<int:event_id>', methods=['POST'])
@login_required
@admin_required
def approve_materials(event_id):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE project_calendar 
            SET status = 'approved'
            WHERE id = ?
        """, (event_id,))
        
        conn.commit()
        conn.close()
        
        flash('Materials request approved!', 'success')
    except Exception as e:
        flash(f'Error approving materials: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))

@app.route('/production_calendar/reject_materials/<int:event_id>', methods=['POST'])
@login_required
@admin_required
def reject_materials(event_id):
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE project_calendar 
            SET status = 'rejected', materials_needed = NULL
            WHERE id = ?
        """, (event_id,))
        
        conn.commit()
        conn.close()
        
        flash('Materials request rejected!', 'success')
    except Exception as e:
        flash(f'Error rejecting materials: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))


@app.route('/projects_management')
@login_required
@admin_required
def projects_management():
    """Project management page for admin users"""
    projects = Project.query.all()
    return render_template('projects.html', projects=projects)

@app.route('/add_project', methods=['GET', 'POST'])
@login_required
@admin_required
def add_project():
    """Add a new project"""
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        client_name = request.form.get('client_name')
        address = request.form.get('address')
        start_date = request.form.get('start_date')
        estimated_end_date = request.form.get('estimated_end_date')
        budget = request.form.get('budget')
        notes = request.form.get('notes')
        
        # Convert date strings to date objects
        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        end_date_obj = datetime.strptime(estimated_end_date, '%Y-%m-%d').date() if estimated_end_date else None
        
        new_project = Project(
            project_name=project_name,
            client_name=client_name,
            address=address,
            start_date=start_date_obj,
            estimated_end_date=end_date_obj,
            budget=float(budget) if budget else 0.0,
            notes=notes,
            status='active'
        )
        
        db.session.add(new_project)
        db.session.commit()
        flash('Project added successfully', 'success')
        return redirect(url_for('projects_management'))
    
    return render_template('add_project.html')

@app.route('/edit_project/<int:project_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_project(project_id):
    """Edit an existing project"""
    project = Project.query.get_or_404(project_id)
    
    if request.method == 'POST':
        project.project_name = request.form.get('project_name')
        project.client_name = request.form.get('client_name')
        project.address = request.form.get('address')
        
        # Handle dates
        start_date = request.form.get('start_date')
        estimated_end_date = request.form.get('estimated_end_date')
        project.start_date = datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
        project.estimated_end_date = datetime.strptime(estimated_end_date, '%Y-%m-%d').date() if estimated_end_date else None
        
        project.budget = float(request.form.get('budget')) if request.form.get('budget') else 0.0
        project.notes = request.form.get('notes')
        project.status = request.form.get('status')
        
        db.session.commit()
        flash('Project updated successfully', 'success')
        return redirect(url_for('projects_management'))
    
    return render_template('edit_project.html', project=project)

@app.route('/delete_project/<int:project_id>', methods=['POST'])
@login_required
@admin_required
def delete_project(project_id):
    """Delete a project"""
    project = Project.query.get_or_404(project_id)
    db.session.delete(project)
    db.session.commit()
    flash('Project deleted successfully', 'success')
    return redirect(url_for('projects_management'))



# ... all your existing code ...

# Remove or modify the development server section
# This prevents accidentally running in development mode on production

# Only run development server if executed directly AND in development
if __name__ == '__main__':
    if os.environ.get('FLASK_ENV') != 'production':
        with app.app_context():
            app.run(debug=True, host='0.0.0.0', port=5000)
    else:
        print("Use Gunicorn or a proper WSGI server for production")
        print("Example: gunicorn --bind 0.0.0.0:5000 wsgi:app")
